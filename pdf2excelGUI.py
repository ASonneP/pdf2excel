import tkinter as tk
from tkinterdnd2 import DND_FILES, TkinterDnD
from tkinter import messagebox
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Alignment

def extract_text_and_create_excel(pdf_path):
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            wb = Workbook()
            ws = wb.active
            middle_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set headers
            headers = ['Page', 'Content']
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).alignment = middle_alignment
            
            # Set column width
            ws.column_dimensions['B'].width = 150

            # Extract text from PDF and write to Excel
            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                page_content = page.extract_text() or "No text found"
                row = [page_num + 1, page_content]
                ws.append(row)
                for col in range(1, len(row) + 1):
                    ws.cell(row=ws.max_row, column=col).alignment = middle_alignment
            
            # Save the Excel file
            excel_path = pdf_path.replace('.pdf', '_content.xlsx')
            wb.save(excel_path)
            messagebox.showinfo("Success", f"Excel file created: {excel_path}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def drop(event):
    file_path = event.data
    if file_path.endswith('.pdf'):
        extract_text_and_create_excel(file_path)
    else:
        messagebox.showwarning("Invalid File", "Please drop a PDF file.")

# Setup the GUI
root = TkinterDnD.Tk()
root.title('PDF to Excel Converter')
root.geometry('400x250')

label = tk.Label(root, text="Drag and drop a PDF file here", pady=20, padx=20)
label.pack(expand=True, fill=tk.BOTH)

# Enable drag and drop
label.drop_target_register(DND_FILES)
label.dnd_bind('<<Drop>>', drop)

root.mainloop()
